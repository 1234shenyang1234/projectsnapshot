from flask import Flask, render_template, request, redirect, url_for, send_from_directory, jsonify, session, g
import os
import datetime
import sqlite3
from PIL import Image, ImageDraw, ImageFont
import base64
from io import BytesIO
import json
from flask import send_file
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import uuid
import hashlib
import socket
import platform
import zipfile
import shutil
import tempfile

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/photos'
app.config['SECRET_KEY'] = 'your-secret-key'

# 确保上传文件夹存在
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 初始化数据库
def init_db():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT,
        timestamp TEXT,
        latitude REAL,
        longitude REAL,
        project_name TEXT,
        project_category TEXT,
        project_town TEXT,
        defect_type TEXT,
        project_status TEXT DEFAULT '计划',
        created_at TEXT,
        updated_at TEXT,
        deadline TEXT,
        remark TEXT
    )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        show_datetime BOOLEAN,
        show_location BOOLEAN,
        design_unit TEXT,
        design_person TEXT,
        show_latlon BOOLEAN,
        show_altitude BOOLEAN,
        remark TEXT,
        watermark_position TEXT DEFAULT '左下角',
        manual_watermark BOOLEAN DEFAULT 0,
        font_family TEXT DEFAULT 'Arial',
        font_size INTEGER DEFAULT 16,
        photo_save_location TEXT DEFAULT '存储卡->PhotoToolData文件夹下'
    )
    ''')
    
    # 创建表存储不同模块的可选条目
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS module_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        module TEXT NOT NULL,
        name TEXT NOT NULL,
        created_at TEXT,
        updated_at TEXT
    )
    ''')
    
    # 创建注册信息表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS registration (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        machine_code TEXT NOT NULL,
        register_code TEXT,
        is_registered BOOLEAN DEFAULT 0,
        created_at TEXT,
        updated_at TEXT
    )
    ''')
    
    # 插入默认设置
    cursor.execute('SELECT * FROM settings')
    if not cursor.fetchone():
        cursor.execute('''
        INSERT INTO settings (show_datetime, show_location, design_unit, design_person, 
                              show_latlon, show_altitude, remark, watermark_position,
                              manual_watermark, font_family, font_size, photo_save_location) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (True, True, '默认设计单位', '默认设计人员', 
              True, True, '', '左下角', 
              0, 'Arial', 16, '存储卡->PhotoToolData文件夹下'))
    
    # 插入默认的模块条目
    cursor.execute('SELECT * FROM module_items')
    if not cursor.fetchone():
        default_items = [
            # 项目名称模块
            ('项目名称', '缺陷隐患照片'),
            ('项目名称', '光伏现场照片'),
            ('项目名称', '老旧表箱照片'),
            ('项目名称', '接户线照片'),
            # 项目分类模块
            ('项目分类', '新建照片'),
            ('项目分类', '整改照片'),
            ('项目分类', '验收照片'),
            # 项目所属乡镇模块
            ('项目所属乡镇', '城镇'),
            ('项目所属乡镇', '农村'),
            ('项目所属乡镇', '工业园区'),
            # 缺陷类型模块
            ('缺陷类型', '轻微缺陷'),
            ('缺陷类型', '一般缺陷'),
            ('缺陷类型', '严重缺陷'),
            # 项目状态模块
            ('项目状态', '计划'),
            ('项目状态', '进行中'),
            ('项目状态', '完成'),
            ('项目状态', '延期')
        ]
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for module, name in default_items:
            cursor.execute('''
            INSERT INTO module_items (module, name, created_at, updated_at) 
            VALUES (?, ?, ?, ?)
            ''', (module, name, now, now))
    
    conn.commit()
    conn.close()

# 生成机器码
def generate_machine_code():
    # 获取设备信息
    hostname = socket.gethostname()
    platform_info = platform.system() + platform.version()
    
    # 生成唯一标识
    unique_info = f"{hostname}-{platform_info}"
    
    # 使用MD5哈希生成机器码
    machine_code = hashlib.md5(unique_info.encode()).hexdigest().upper()
    
    # 格式化机器码，每4位一组，用-分隔
    formatted_code = '-'.join([machine_code[i:i+4] for i in range(0, len(machine_code), 4)])
    
    return formatted_code

# 检查注册状态
def is_registered():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT is_registered FROM registration WHERE id = 1')
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return result[0] == 1
    return False

# 生成注册码（注册机使用）
def generate_register_code(machine_code):
    # 简单的注册码生成算法，实际应用中应使用更复杂的算法
    # 移除机器码中的分隔符
    clean_machine = machine_code.replace('-', '')
    
    # 生成注册码
    register_code = hashlib.sha256((clean_machine + 'your-secret-key').encode()).hexdigest().upper()
    
    # 格式化注册码，每5位一组，用-分隔
    formatted_register = '-'.join([register_code[i:i+5] for i in range(0, len(register_code), 5)])
    
    return formatted_register

# 验证注册码
def verify_register_code(machine_code, register_code):
    # 生成预期的注册码
    expected_code = generate_register_code(machine_code)
    
    # 验证注册码
    return register_code == expected_code

# 注册状态检查装饰器
def register_required(f):
    def decorated_function(*args, **kwargs):
        # 检查注册状态
        if not is_registered():
            # 允许访问的页面：注册页面、登录页面（如果有的话）
            if request.path in ['/register', '/verify_register']:
                return f(*args, **kwargs)
            return redirect('/register')
        return f(*args, **kwargs)
    # 设置装饰后的函数名称为原始函数名称，避免端点冲突
    decorated_function.__name__ = f.__name__
    decorated_function.__module__ = f.__module__
    return decorated_function

@app.route('/')
def index():
    # 检查注册状态，如果未注册则重定向到注册页面
    if not is_registered():
        return redirect('/register')
    return redirect(url_for('main'))

@app.route('/main')
@register_required
def main():
    return render_template('main.html')

@app.route('/camera')
@register_required
def camera():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings WHERE id = 1')
    settings = cursor.fetchone()
    conn.close()
    return render_template('camera.html', settings=settings)

@app.route('/album')
@register_required
def album():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM photos ORDER BY timestamp DESC')
    photos = cursor.fetchall()
    conn.close()
    return render_template('album.html', photos=photos)

@app.route('/profile')
@register_required
def profile():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings WHERE id = 1')
    settings = cursor.fetchone()
    conn.close()
    return render_template('profile.html', settings=settings)

@app.route('/部位选择')
@app.route('/部位选择/<module_name>')
@register_required
def part_selection(module_name=None):
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    # 获取指定模块的条目
    if module_name:
        cursor.execute('SELECT * FROM module_items WHERE module = ?', (module_name,))
    else:
        cursor.execute('SELECT * FROM module_items')
    
    module_items = cursor.fetchall()
    conn.close()
    
    return render_template('部位选择.html', module_name=module_name, module_items=module_items)

# 模板页面路由
@app.route('/template')
@register_required
def template():
    return render_template('template.html')

# 定位页面路由
@app.route('/location')
@register_required
def location():
    return render_template('location.html')

# 图片保存位置页面路由
@app.route('/save_location')
@register_required
def save_location_page():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings WHERE id = 1')
    settings = cursor.fetchone()
    conn.close()
    return render_template('save_location.html', settings=settings)

# 保存图片保存位置设置
@app.route('/save_location', methods=['POST'])
@register_required
def save_location():
    photo_save_location = request.form.get('photo_save_location', '存储卡->PhotoToolData文件夹下')
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('''
    UPDATE settings SET photo_save_location = ? WHERE id = 1
    ''', (photo_save_location,))
    conn.commit()
    conn.close()
    
    return redirect(url_for('save_location_page'))

@app.route('/settings')
@register_required
def settings():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings WHERE id = 1')
    settings = cursor.fetchone()
    conn.close()
    return render_template('settings.html', settings=settings)

@app.route('/save_settings', methods=['POST'])
@register_required
def save_settings():
    show_datetime = request.form.get('show_datetime') == 'on'
    show_location = request.form.get('show_location') == 'on'
    design_unit = request.form.get('design_unit', '默认设计单位')
    design_person = request.form.get('design_person', '默认设计人员')
    show_latlon = request.form.get('show_latlon') == 'on'
    show_altitude = request.form.get('show_altitude') == 'on'
    remark = request.form.get('remark', '')
    watermark_position = request.form.get('watermark_position', '左下角')
    manual_watermark = request.form.get('manual_watermark') == 'on'
    font_family = request.form.get('font_family', 'Arial')
    font_size = int(request.form.get('font_size', 16))
    photo_save_location = request.form.get('photo_save_location', '存储卡->PhotoToolData文件夹下')
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('''
    UPDATE settings SET show_datetime = ?, show_location = ?, design_unit = ?, 
                   design_person = ?, show_latlon = ?, show_altitude = ?, 
                   remark = ?, watermark_position = ?, manual_watermark = ?, 
                   font_family = ?, font_size = ?, photo_save_location = ? 
    WHERE id = 1
    ''', (show_datetime, show_location, design_unit, design_person, 
          show_latlon, show_altitude, remark, watermark_position, 
          manual_watermark, font_family, font_size, photo_save_location))
    conn.commit()
    conn.close()
    
    return redirect(url_for('settings'))

@app.route('/record/<int:record_id>')
@register_required
def record_detail(record_id):
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM photos WHERE id = ?', (record_id,))
    photo = cursor.fetchone()
    conn.close()
    return render_template('record_detail.html', photo=photo)

@app.route('/save_record/<int:record_id>', methods=['POST'])
@register_required
def save_record(record_id):
    project_name = request.form.get('project_name')
    project_category = request.form.get('project_category')
    project_town = request.form.get('project_town')
    defect_type = request.form.get('defect_type')
    project_status = request.form.get('project_status')
    created_at = request.form.get('created_at')
    deadline = request.form.get('deadline')
    remark = request.form.get('remark')
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('''
    UPDATE photos SET project_name = ?, project_category = ?, project_town = ?, 
                   defect_type = ?, project_status = ?, created_at = ?, 
                   deadline = ?, remark = ?, updated_at = ?
    WHERE id = ?
    ''', (project_name, project_category, project_town, defect_type, 
          project_status, created_at, deadline, remark, 
          datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), record_id))
    conn.commit()
    conn.close()
    
    return redirect(url_for('record_detail', record_id=record_id))

# 保存新的模块条目
@app.route('/save_module_item', methods=['POST'])
@register_required
def save_module_item():
    import json
    data = request.get_json()
    module = data.get('module')
    name = data.get('name')
    
    if module and name:
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
        INSERT INTO module_items (module, name, created_at, updated_at) 
        VALUES (?, ?, ?, ?)
        ''', (module, name, now, now))
        conn.commit()
        conn.close()
        return {'success': True}
    return {'success': False}

# 更新模块条目
@app.route('/update_module_item', methods=['POST'])
@register_required
def update_module_item():
    import json
    data = request.get_json()
    item_id = data.get('id')
    name = data.get('name')
    
    if item_id and name:
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
        UPDATE module_items SET name = ?, updated_at = ? WHERE id = ?
        ''', (name, now, item_id))
        conn.commit()
        conn.close()
        return {'success': True}
    return {'success': False}

# 删除模块条目
@app.route('/delete_module_item', methods=['POST'])
@register_required
def delete_module_item():
    import json
    data = request.get_json()
    item_id = data.get('id')
    
    if item_id:
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        cursor.execute('DELETE FROM module_items WHERE id = ?', (item_id,))
        conn.commit()
        conn.close()
        return {'success': True}
    return {'success': False}

# 删除照片
@app.route('/delete_photos', methods=['POST'])
@register_required
def delete_photos():
    import json
    data = request.get_json()
    ids = data.get('ids', [])
    
    if ids:
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        # 构建IN语句
        placeholders = ','.join(['?' for _ in ids])
        cursor.execute(f'DELETE FROM photos WHERE id IN ({placeholders})', ids)
        conn.commit()
        conn.close()
        return {'success': True}
    return {'success': False}

# 获取模块条目
@app.route('/get_module_items')
@register_required
def get_module_items():
    module = request.args.get('module')
    if not module:
        return jsonify([])
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM module_items WHERE module = ?', (module,))
    items = cursor.fetchall()
    conn.close()
    
    # 转换为JSON格式
    result = []
    for item in items:
        result.append({
            'id': item[0],
            'module': item[1],
            'name': item[2],
            'created_at': item[3],
            'updated_at': item[4]
        })
    
    return jsonify(result)

# 批量保存模块条目
@app.route('/batch_save_module_items', methods=['POST'])
@register_required
def batch_save_module_items():
    data = request.get_json()
    module = data.get('module')
    names = data.get('names', [])
    
    if not module or not names:
        return jsonify({'success': False})
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    added = 0
    
    for name in names:
        name = name.strip()
        if name:
            cursor.execute('''
            INSERT INTO module_items (module, name, created_at, updated_at) 
            VALUES (?, ?, ?, ?)
            ''', (module, name, now, now))
            added += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'added': added})

# 批量删除模块条目
@app.route('/batch_delete_module_items', methods=['POST'])
@register_required
def batch_delete_module_items():
    data = request.get_json()
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'success': False})
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    # 构建IN语句
    placeholders = ','.join(['?' for _ in ids])
    cursor.execute(f'DELETE FROM module_items WHERE id IN ({placeholders})', ids)
    deleted = cursor.rowcount
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'deleted': deleted})

# 导入Excel
@app.route('/import_excel', methods=['POST'])
@register_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有文件上传'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        processed = 0
        
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 支持的模块列表
        modules = ['项目名称', '项目分类', '项目所属乡镇', '缺陷类型', '项目状态']
        
        for module in modules:
            if module in wb.sheetnames:
                ws = wb[module]
                # 读取数据行（跳过表头）
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[1].strip():
                        name = row[1].strip()
                        cursor.execute('''
                        INSERT INTO module_items (module, name, created_at, updated_at) 
                        VALUES (?, ?, ?, ?)
                        ''', (module, name, now, now))
                        processed += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'processed': processed})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# 导出Excel
@app.route('/export_excel')
@register_required
def export_excel():
    # 创建Excel工作簿
    wb = Workbook()
    
    # 支持的模块列表
    modules = ['项目名称', '项目分类', '项目所属乡镇', '缺陷类型', '项目状态']
    
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    for module in modules:
        # 创建工作表
        ws = wb.create_sheet(module)
        # 设置表头
        ws.append(['序号', '条目名称', '备注'])
        
        # 获取数据
        cursor.execute('SELECT * FROM module_items WHERE module = ?', (module,))
        items = cursor.fetchall()
        
        # 写入数据
        for i, item in enumerate(items, 1):
            ws.append([i, item[2], ''])
    
    conn.close()
    
    # 删除默认的工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name='模板数据.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 下载模板
@app.route('/download_template')
@register_required
def download_template():
    # 创建模板工作簿
    wb = Workbook()
    
    # 支持的模块列表
    modules = ['项目名称', '项目分类', '项目所属乡镇', '缺陷类型', '项目状态']
    
    for module in modules:
        # 创建工作表
        ws = wb.create_sheet(module)
        # 设置表头
        ws.append(['序号', '条目名称', '备注'])
        # 添加示例数据
        ws.append([1, '示例1', ''])
        ws.append([2, '示例2', ''])
    
    # 删除默认的工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name='模板下载.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 导出选中照片信息
@app.route('/export_photos')
@register_required
def export_photos():
    # 获取选中的照片ID列表
    ids_param = request.args.get('ids')
    if not ids_param:
        return jsonify({'success': False, 'message': '没有选择照片'})
    
    # 转换为整数列表
    ids = [int(id_str) for id_str in ids_param.split(',')]
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 创建照片信息工作表
    ws = wb.active
    ws.title = '照片信息'
    
    # 设置表头
    headers = ['ID', '文件名', '拍摄时间', '纬度', '经度', '项目名称', '项目分类', 
              '项目所属乡镇', '缺陷类型', '项目状态', '创建时间', '更新时间', '截止日期', '备注']
    ws.append(headers)
    
    # 查询选中的照片信息
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    # 构建IN语句
    placeholders = ','.join(['?' for _ in ids])
    cursor.execute(f'SELECT * FROM photos WHERE id IN ({placeholders})', ids)
    photos = cursor.fetchall()
    
    # 写入数据行
    for photo in photos:
        ws.append(photo)
    
    conn.close()
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回Excel文件
    return send_file(buffer, as_attachment=True, download_name='照片信息.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 导入照片信息
@app.route('/import_photos', methods=['POST'])
@register_required
def import_photos():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有文件上传'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        # 获取第一个工作表
        ws = wb.active
        
        conn = sqlite3.connect('project_snapshot.db')
        cursor = conn.cursor()
        updated = 0
        
        # 读取数据行（跳过表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 确保行数据完整
            if len(row) < 14:
                continue
            
            # 获取ID
            photo_id = row[0]
            if not photo_id:
                continue
            
            # 更新照片信息
            cursor.execute('''
            UPDATE photos SET filename = ?, timestamp = ?, latitude = ?, longitude = ?, 
                   project_name = ?, project_category = ?, project_town = ?, 
                   defect_type = ?, project_status = ?, created_at = ?, 
                   updated_at = ?, deadline = ?, remark = ? 
            WHERE id = ?
            ''', (row[1], row[2], row[3], row[4], row[5], row[6], row[7], 
                  row[8], row[9], row[10], row[11], row[12], row[13], photo_id))
            
            if cursor.rowcount > 0:
                updated += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# 获取点位数据
@app.route('/get_points')
@register_required
def get_points():
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    # 查询所有带有经纬度的照片
    cursor.execute('SELECT * FROM photos WHERE latitude IS NOT NULL AND longitude IS NOT NULL')
    points = cursor.fetchall()
    conn.close()
    
    # 转换为JSON格式
    result = []
    for point in points:
        result.append({
            'id': point[0],
            'filename': point[1],
            'timestamp': point[2],
            'latitude': point[3],
            'longitude': point[4],
            'project_name': point[5],
            'project_category': point[6],
            'project_town': point[7],
            'defect_type': point[8],
            'project_status': point[9],
            'created_at': point[10],
            'updated_at': point[11]
        })
    
    return jsonify(result)

@app.route('/capture', methods=['POST'])
@register_required
def capture():
    import json
    
    # 添加调试信息
    print("=== 收到capture请求 ===")
    print(f"请求表单数据: {request.form}")
    
    images_json = request.form.get('images')
    latitude = request.form.get('latitude', 0.0)
    longitude = request.form.get('longitude', 0.0)
    
    # 获取设置
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings WHERE id = 1')
    settings = cursor.fetchone()
    conn.close()
    
    # 直接创建一个记录，确保跳转到详情页面
    print("=== 直接创建记录，确保跳转到详情页面 ===")
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    try:
        # 尝试使用简化的插入语句，只使用基本列
        cursor.execute('''
        INSERT INTO photos (filename, timestamp, latitude, longitude, project_name)
        VALUES (?, ?, ?, ?, ?)
        ''', ('default.jpg', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 
              latitude, longitude, settings[3]))
        conn.commit()
        record_id = cursor.lastrowid
        conn.close()
        print(f"=== 使用基本列插入成功，ID: {record_id} ===")
    except Exception as e:
        print(f"=== 基本列插入失败: {e} ===")
        # 尝试获取表结构，然后根据实际表结构插入
        cursor.execute('PRAGMA table_info(photos)')
        columns = [row[1] for row in cursor.fetchall()]
        print(f"=== 当前表结构: {columns} ===")
        
        # 使用表中实际存在的列插入
        basic_columns = ['filename', 'timestamp', 'latitude', 'longitude', 'project_name']
        available_columns = [col for col in basic_columns if col in columns]
        
        if available_columns:
            print(f"=== 使用可用列: {available_columns} ===")
            insert_sql = f"INSERT INTO photos ({', '.join(available_columns)}) VALUES ({', '.join(['?'] * len(available_columns))})"
            insert_values = []
            for col in available_columns:
                if col == 'filename':
                    insert_values.append('default.jpg')
                elif col == 'timestamp':
                    insert_values.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                elif col == 'latitude':
                    insert_values.append(latitude)
                elif col == 'longitude':
                    insert_values.append(longitude)
                elif col == 'project_name':
                    insert_values.append(settings[3])
            
            cursor.execute(insert_sql, insert_values)
            conn.commit()
            record_id = cursor.lastrowid
            conn.close()
            print(f"=== 使用可用列插入成功，ID: {record_id} ===")
        else:
            conn.close()
            print("=== 无法插入记录，表结构不匹配 ===")
            # 如果无法插入，跳转到相册页面
            return redirect(url_for('album'))
    
    print(f"=== 成功创建记录，ID: {record_id}，跳转到详情页面 ===")
    return redirect(url_for('record_detail', record_id=record_id))

@app.route('/static/photos/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 打包选中照片
@app.route('/package_photos')
@register_required
def package_photos():
    # 获取选中的照片ID列表
    ids_param = request.args.get('ids')
    if not ids_param:
        return jsonify({'success': False, 'message': '没有选择照片'})
    
    # 转换为整数列表
    ids = [int(id_str) for id_str in ids_param.split(',')]
    
    # 获取打包格式
    format = request.args.get('format', 'zip').lower()
    if format not in ['zip', '7z']:
        format = 'zip'
    
    # 查询选中的照片信息
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    # 构建IN语句
    placeholders = ','.join(['?' for _ in ids])
    cursor.execute(f'SELECT * FROM photos WHERE id IN ({placeholders})', ids)
    photos = cursor.fetchall()
    conn.close()
    
    if not photos:
        return jsonify({'success': False, 'message': '未找到选中的照片'})
    
    try:
        # 创建临时目录
        with tempfile.TemporaryDirectory() as temp_dir:
            # 确定打包文件名
            project_name = photos[0][5] if photos and photos[0][5] else '项目随手拍'
            package_name = f'{project_name}.{format}'
            package_path = os.path.join(temp_dir, package_name)
            
            if format == 'zip':
                # 创建zip压缩文件
                with zipfile.ZipFile(package_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for photo in photos:
                        # 获取照片文件路径
                        photo_filename = photo[1]
                        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_filename)
                        
                        # 检查照片文件是否存在
                        if os.path.exists(photo_path):
                            # 将照片添加到压缩文件
                            zipf.write(photo_path, photo_filename)
            elif format == '7z':
                # 尝试使用py7zr创建7z压缩文件
                try:
                    import py7zr
                    with py7zr.SevenZipFile(package_path, 'w') as z:
                        for photo in photos:
                            photo_filename = photo[1]
                            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_filename)
                            if os.path.exists(photo_path):
                                z.write(photo_path, photo_filename)
                except ImportError:
                    # 如果py7zr未安装，降级使用zip格式
                    package_name = f'{project_name}.zip'
                    package_path = os.path.join(temp_dir, package_name)
                    with zipfile.ZipFile(package_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for photo in photos:
                            photo_filename = photo[1]
                            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_filename)
                            if os.path.exists(photo_path):
                                zipf.write(photo_path, photo_filename)
            
            # 返回压缩文件
            return send_file(package_path, as_attachment=True, download_name=package_name)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# 注册页面路由
@app.route('/register')
def register():
    # 生成机器码
    machine_code = generate_machine_code()
    
    # 保存机器码到数据库
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    
    # 检查是否已经存在注册信息
    cursor.execute('SELECT id FROM registration WHERE id = 1')
    result = cursor.fetchone()
    
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if result:
        # 更新机器码
        cursor.execute('''
        UPDATE registration SET machine_code = ?, updated_at = ? WHERE id = 1
        ''', (machine_code, now))
    else:
        # 插入新的注册信息
        cursor.execute('''
        INSERT INTO registration (machine_code, is_registered, created_at, updated_at)
        VALUES (?, ?, ?, ?)
        ''', (machine_code, 0, now, now))
    
    conn.commit()
    conn.close()
    
    return render_template('register.html', machine_code=machine_code)

# 验证注册码路由
@app.route('/verify_register', methods=['POST'])
def verify_register():
    # 获取注册码
    data = request.get_json()
    register_code = data.get('register_code')
    
    if not register_code:
        return jsonify({'success': False, 'message': '请输入注册码'})
    
    # 获取机器码
    conn = sqlite3.connect('project_snapshot.db')
    cursor = conn.cursor()
    cursor.execute('SELECT machine_code FROM registration WHERE id = 1')
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return jsonify({'success': False, 'message': '无法获取机器码'})
    
    machine_code = result[0]
    
    # 验证注册码
    if verify_register_code(machine_code, register_code):
        # 更新注册状态
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
        UPDATE registration SET register_code = ?, is_registered = 1, updated_at = ? WHERE id = 1
        ''', (register_code, now))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '注册成功'})
    else:
        conn.close()
        return jsonify({'success': False, 'message': '注册码无效'})

if __name__ == '__main__':
    init_db()
    app.run(debug=True)
